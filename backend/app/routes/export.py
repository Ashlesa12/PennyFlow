import csv
import io
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends
from fastapi.responses import Response
from sqlalchemy.orm import Session, joinedload
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models import Expense, User
from app.security import get_current_user
from app.services.expense_filters import apply_expense_filters

router = APIRouter(
    prefix="/export",
    tags=["Export"]
)

CSV_MEDIA_TYPE = "text/csv; charset=utf-8"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

COLUMNS = ["ID", "Date", "Title", "Amount", "Category"]


def _build_filename(ext: str) -> str:
    return f"expenses_{date.today().isoformat()}.{ext}"


def _fetch_expenses(
    db: Session,
    current_user: User,
    search: Optional[str],
    category_id: Optional[int],
    min_amount: Optional[float],
    max_amount: Optional[float],
    start_date: Optional[date],
    end_date: Optional[date],
    month: Optional[int] = None,
    year: Optional[int] = None,
):

    query = db.query(Expense).filter(
        Expense.user_id == current_user.id
    )

    query = apply_expense_filters(
        query,
        search=search,
        category_id=category_id,
        min_amount=min_amount,
        max_amount=max_amount,
        start_date=start_date,
        end_date=end_date,
        month=month,
        year=year,
    )

    return query.options(joinedload(Expense.category)).order_by(
        Expense.expense_date.desc()
    ).all()


def _category_name(expense: Expense) -> str:
    return expense.category.name if expense.category else ""


def _build_csv(expenses):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(COLUMNS)

    for expense in expenses:
        writer.writerow([
            expense.id,
            expense.expense_date.isoformat(),
            expense.title,
            str(expense.amount),
            _category_name(expense),
        ])

    return output.getvalue().encode("utf-8-sig")


def _build_xlsx(expenses) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Expenses"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")

    sheet.append(COLUMNS)

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for expense in expenses:
        sheet.append([
            expense.id,
            expense.expense_date.isoformat(),
            expense.title,
            float(expense.amount),
            _category_name(expense),
        ])

    widths = [8, 12, 32, 12, 20]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"

    output = io.BytesIO()
    workbook.save(output)

    return output.getvalue()


@router.get("/expenses.csv")
def export_expenses_csv(
    search: Optional[str] = None,
    category_id: Optional[int] = None,
    min_amount: Optional[float] = None,
    max_amount: Optional[float] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):

    expenses = _fetch_expenses(
        db,
        current_user,
        search=search,
        category_id=category_id,
        min_amount=min_amount,
        max_amount=max_amount,
        start_date=start_date,
        end_date=end_date,
        month=month,
        year=year,
    )

    return Response(
        content=_build_csv(expenses),
        media_type=CSV_MEDIA_TYPE,
        headers={
            "Content-Disposition": f'attachment; filename="{_build_filename("csv")}"'
        },
    )


@router.get("/expenses.xlsx")
def export_expenses_xlsx(
    search: Optional[str] = None,
    category_id: Optional[int] = None,
    min_amount: Optional[float] = None,
    max_amount: Optional[float] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):

    expenses = _fetch_expenses(
        db,
        current_user,
        search=search,
        category_id=category_id,
        min_amount=min_amount,
        max_amount=max_amount,
        start_date=start_date,
        end_date=end_date,
        month=month,
        year=year,
    )

    return Response(
        content=_build_xlsx(expenses),
        media_type=XLSX_MEDIA_TYPE,
        headers={
            "Content-Disposition": f'attachment; filename="{_build_filename("xlsx")}"'
        },
    )
